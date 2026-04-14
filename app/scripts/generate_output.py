"""
Output generation for LP field mapping results.

Generates:
- Formatted Excel files with mapping results
- JSON config files for direct LOS system integration
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_mapping_excel(
    mappings: list[dict],
    output_path: str,
    client_name: str = "",
    process_name: str = "",
    missing_mandatory: Optional[list[dict]] = None,
) -> None:
    """
    Generate a formatted Excel file with the mapping results.

    Args:
        mappings: List of mapping dicts with keys:
                  partner_field, column_category, entity, matched_excel_key,
                  json_key, confidence, match_type, reasoning, needs_review
        output_path: Path where Excel file will be written
        client_name: Optional client name for metadata
        process_name: Optional process name for metadata

    Sheet layout:
    - Sheet 1: all mappings
    - Sheet 2: summary
    - Sheet 3: review-only mappings
    """
    if not HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for Excel generation. "
            "Install with: pip install openpyxl"
        )

    wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    review_mappings = [m for m in mappings if m.get("needs_review", False)]

    # Sheet 1: Field Mapping — all mappings
    ws_mapping = wb.create_sheet("Field Mapping", 0)
    _populate_mapping_sheet(ws_mapping, mappings)

    # Sheet 2: Summary — uses all mappings for accurate stats
    ws_summary = wb.create_sheet("Summary", 1)
    _populate_summary_sheet(
        ws_summary,
        mappings,
        client_name,
        process_name,
        missing_mandatory=missing_mandatory or [],
    )

    # Sheet 3: For Review — needs_review mappings only
    if review_mappings:
        ws_review = wb.create_sheet("For Review", 2)
        _populate_mapping_sheet(ws_review, review_mappings)

    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path_obj)

    print(f"Excel mapping file generated: {output_path}")


def _populate_mapping_sheet(ws, mappings: list[dict]) -> None:
    """
    Populate a mapping sheet (Field Mapping or For Review) with data and formatting.

    Args:
        ws: Openpyxl worksheet object
        mappings: List of mapping dicts
    """
    headers = [
        "Partner Field",
        "Category",
        "Entity",
        "Mapped Excel Key",
        "JSON Key",
        "Confidence",
        "Match Type",
        "Reasoning",
        "Previous Mapping Reason",
        "LLM Change Reason",
        "LLM Param Bucket Reason",
        "Needs Review"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    REVIEW_BG = "FFFFCC"   # pale yellow for needs_review rows

    for row_num, mapping in enumerate(mappings, 2):
        needs_review = mapping.get("needs_review", False)

        cells_data = [
            mapping.get("partner_field", ""),
            mapping.get("column_category", ""),
            mapping.get("entity", ""),
            mapping.get("matched_excel_key", ""),
            mapping.get("json_key", ""),
            mapping.get("confidence", 0.0),
            mapping.get("match_type", ""),
            mapping.get("reasoning", ""),
            mapping.get("previous_mapping_reason", ""),
            mapping.get("llm_change_reason", ""),
            mapping.get("llm_param_bucket_reason", ""),
            "YES" if needs_review else "NO"
        ]

        for col_num, value in enumerate(cells_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            # Confidence column (col 6): colour by band
            if col_num == 6:
                confidence = value if isinstance(value, (int, float)) else 0.0
                if confidence >= 0.90:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif confidence >= 0.80:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif confidence >= 0.70:
                    cell.fill = PatternFill(start_color="FED8B1", end_color="FED8B1", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

            # FIX: Highlight needs_review rows in pale yellow — but skip column 6
            # so the confidence colour is preserved and not overwritten.
            # Previously the whole row including col 6 was overwritten, masking the
            # confidence colour. Now col 6 always keeps its confidence band colour.
            elif needs_review:
                cell.fill = PatternFill(
                    start_color=REVIEW_BG, end_color=REVIEW_BG, fill_type="solid"
                )

    column_widths = {
        "A": 25,  # Partner Field
        "B": 15,  # Category
        "C": 15,  # Entity
        "D": 18,  # Mapped Excel Key
        "E": 35,  # JSON Key
        "F": 12,  # Confidence
        "G": 18,  # Match Type
        "H": 40,  # Reasoning
        "I": 45,  # Previous Mapping Reason
        "J": 45,  # LLM Change Reason
        "K": 45,  # LLM Param Bucket Reason
        "L": 12,  # Needs Review
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(mappings) + 1}"


def _populate_summary_sheet(
    ws,
    mappings: list[dict],
    client_name: str = "",
    process_name: str = "",
    missing_mandatory: Optional[list[dict]] = None,
) -> None:
    """
    Populate the Summary sheet with statistics.

    FIX: Added separate counts for confirmed vs needs_review fields so the
    summary accurately reflects what appears in each sheet.
    """
    row = 1

    if client_name:
        ws.cell(row=row, column=1).value = "Client:"
        ws.cell(row=row, column=2).value = client_name
        row += 1

    if process_name:
        ws.cell(row=row, column=1).value = "Process:"
        ws.cell(row=row, column=2).value = process_name
        row += 1

    ws.cell(row=row, column=1).value = "Generated:"
    ws.cell(row=row, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    row += 2

    # FIX: Show grand total + breakdown by sheet
    review_mappings = [m for m in mappings if m.get("needs_review", False)]
    clean_mappings  = [m for m in mappings if not m.get("needs_review", False)]

    ws.cell(row=row, column=1).value = "Total Fields Mapped:"
    ws.cell(row=row, column=2).value = len(mappings)
    ws.cell(row=row, column=2).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1).value = "  — Confirmed (Field Mapping sheet):"
    ws.cell(row=row, column=2).value = len(clean_mappings)
    row += 1

    ws.cell(row=row, column=1).value = "  — Needs Review (For Review sheet):"
    ws.cell(row=row, column=2).value = len(review_mappings)
    row += 2

    # Breakdown by match type
    ws.cell(row=row, column=1).value = "Breakdown by Match Type:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    match_type_counts = defaultdict(int)
    for mapping in mappings:
        match_type = mapping.get("match_type", "unknown")
        match_type_counts[match_type] += 1

    for match_type, count in sorted(match_type_counts.items()):
        ws.cell(row=row, column=2).value = match_type
        ws.cell(row=row, column=3).value = count
        row += 1

    row += 1

    # Breakdown by entity
    ws.cell(row=row, column=1).value = "Breakdown by Entity:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    entity_counts = defaultdict(int)
    for mapping in mappings:
        entity = mapping.get("entity", "unknown")
        entity_counts[entity] += 1

    for entity, count in sorted(entity_counts.items()):
        ws.cell(row=row, column=2).value = entity
        ws.cell(row=row, column=3).value = count
        row += 1

    row += 1

    # Fields needing review
    ws.cell(row=row, column=1).value = "Fields Needing Review:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = len(review_mappings)
    ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
    row += 1

    if review_mappings:
        ws.cell(row=row, column=2).value = "Fields:"
        row += 1
        for mapping in review_mappings[:20]:
            ws.cell(row=row, column=3).value = mapping.get("partner_field", "")
            ws.cell(row=row, column=4).value = f"{mapping.get('confidence', 0.0):.2f}"
            row += 1
        if len(review_mappings) > 20:
            ws.cell(row=row, column=3).value = f"... and {len(review_mappings) - 20} more"

    row += 2

    missing_mandatory = missing_mandatory or []
    ws.cell(row=row, column=1).value = "Mandatory Fields Missing:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).value = len(missing_mandatory)
    if missing_mandatory:
        ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
    row += 1

    if missing_mandatory:
        ws.cell(row=row, column=2).value = "Excel Key"
        ws.cell(row=row, column=3).value = "Description"
        row += 1
        for item in missing_mandatory:
            ws.cell(row=row, column=2).value = item.get("excel_key", "")
            ws.cell(row=row, column=3).value = item.get("description", "")
            row += 1

    row += 2

    # Confidence distribution
    ws.cell(row=row, column=1).value = "Confidence Distribution:"
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    confidence_ranges = {
        "0.90-1.00": [m for m in mappings if m.get("confidence", 0.0) >= 0.90],
        "0.80-0.89": [m for m in mappings if 0.80 <= m.get("confidence", 0.0) < 0.90],
        "0.70-0.79": [m for m in mappings if 0.70 <= m.get("confidence", 0.0) < 0.80],
        "0.00-0.69": [m for m in mappings if m.get("confidence", 0.0) < 0.70]
    }

    for range_label, mappings_in_range in confidence_ranges.items():
        ws.cell(row=row, column=2).value = range_label
        ws.cell(row=row, column=3).value = len(mappings_in_range)
        row += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15


def generate_api_config_json(
    mappings: list[dict],
    output_path: str,
    client_name: str = "",
    process_name: str = ""
) -> None:
    """
    Generate a JSON config file for direct LOS system integration.
    Mirrors the generic_excel_upload_definition_fields_dump.csv format.
    """
    config = {
        "metadata": {
            "client_name": client_name,
            "process_name": process_name,
            "generated_at": datetime.now().isoformat(),
            "total_fields": len(mappings),
            "fields_needing_review": sum(1 for m in mappings if m.get("needs_review", False))
        },
        "field_mappings": []
    }

    by_entity = defaultdict(list)
    for mapping in mappings:
        entity = mapping.get("entity", "UNKNOWN")
        by_entity[entity].append(mapping)

    for entity in sorted(by_entity.keys()):
        entity_mappings = by_entity[entity]
        for mapping in entity_mappings:
            api_entry = {
                "partner_field": mapping.get("partner_field", ""),
                "entity": entity,
                "excel_key": mapping.get("matched_excel_key", ""),
                "json_path": mapping.get("json_key", ""),
                "confidence_score": mapping.get("confidence", 0.0),
                "match_strategy": mapping.get("match_type", ""),
                "notes": mapping.get("reasoning", ""),
                "review_required": mapping.get("needs_review", False),
                "category": mapping.get("column_category", None)
            }
            config["field_mappings"].append(api_entry)

    output_path_obj = Path(output_path)
    output_path_obj.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path_obj, "w") as f:
        json.dump(config, f, indent=2)

    print(f"JSON config file generated: {output_path}")


def generate_mapping_summary(mappings: list[dict]) -> dict:
    """Generate a summary of mapping statistics."""
    summary = {
        "total_fields": len(mappings),
        "fields_needing_review": sum(1 for m in mappings if m.get("needs_review", False)),
        "by_match_type": defaultdict(int),
        "by_entity": defaultdict(int),
        "by_confidence_range": {
            "0.90-1.00": 0,
            "0.80-0.89": 0,
            "0.70-0.79": 0,
            "0.00-0.69": 0
        },
        "average_confidence": 0.0
    }

    total_confidence = 0.0

    for mapping in mappings:
        match_type = mapping.get("match_type", "unknown")
        summary["by_match_type"][match_type] += 1

        entity = mapping.get("entity", "unknown")
        summary["by_entity"][entity] += 1

        confidence = mapping.get("confidence", 0.0)
        total_confidence += confidence

        if confidence >= 0.90:
            summary["by_confidence_range"]["0.90-1.00"] += 1
        elif confidence >= 0.80:
            summary["by_confidence_range"]["0.80-0.89"] += 1
        elif confidence >= 0.70:
            summary["by_confidence_range"]["0.70-0.79"] += 1
        else:
            summary["by_confidence_range"]["0.00-0.69"] += 1

    if mappings:
        summary["average_confidence"] = round(total_confidence / len(mappings), 4)

    return summary
