#!/usr/bin/env python3
 #input _parser
import json
import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
 
import pandas as pd
from openpyxl import load_workbook
 
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
 
 
# =========================
# FIELD STRUCTURE
# =========================
class FieldDefinition:
    def __init__(
        self,
        field_name: str,
        column_category: Optional[str] = None,
        data_type: Optional[str] = None,
        is_required: Optional[bool] = None,
        sample_value: Optional[str] = None,
        description: Optional[str] = None,
        source_sheet: Optional[str] = None
    ):
        self.field_name = field_name
        self.column_category = column_category
        self.data_type = data_type
        self.is_required = is_required
        self.sample_value = sample_value
        self.description = description
        self.source_sheet = source_sheet

    def to_dict(self):
        return self.__dict__


# =========================
# IDENTIFIER DETECTION
# =========================
def _looks_like_identifier(val: str) -> bool:
    val = val.rstrip('.,:;')
    words = val.split()
 
    if not words:
        return False
 
    if re.match(r'^[A-Z][A-Za-z0-9_\s]+$', val):
        return True
 
    if re.match(r'^[a-z][a-zA-Z0-9_]+$', val):
        return True
 
    if re.match(r'^[A-Za-z][A-Za-z0-9_]+$', val):
        return True
 
    if 1 <= len(words) <= 6:
        return True
 
    return False


# =========================
# DOTTED-PATH DETECTION
# =========================
def _is_dotted_path(val: str) -> bool:
    """
    Return True if the value looks like a dotted-path field name,
    e.g. 'loanAccount.cbDateTime', 'a.b.c', 'foo.barBaz'.
    A dotted path must have at least one dot with non-empty segments
    on both sides, and no whitespace.
    """
    if ' ' in val:
        return False
    parts = val.split('.')
    if len(parts) < 2:
        return False
    return all(len(p) > 0 for p in parts)


# =========================
# HEADER DETECTION
# =========================

_EXACT_HEADERS = {
    "field names", "field name", "column name", "column names",
    "header", "headers", "s.no", "s.no.", "sr no", "sr. no", "sr.no",
    "input fields", "input field", "data fields", "data field",
    "description", "remarks", "notes", "mandatory", "required",
}

_SINGLE_WORD_HEADERS = {
    "field", "fields", "column", "columns", "header",
    "input", "inputs", "description", "remarks",
}

_LABEL_SUFFIXES = {
    'input', 'inputs', 'fields', 'field', 'field names', 'field name',
    'data', 'details', 'info', 'information', 'section',
}

_GENERIC_WORDS = {
    'client', 'customer', 'applicant', 'loan',
    'coapplicant', 'details', 'info',
}


def _is_header_label(val: str) -> bool:
    if not val:
        return False
 
    stripped = val.strip()
    lower    = stripped.lower()
 
    if lower in _EXACT_HEADERS:
        return True
 
    words = stripped.split()
 
    if len(words) == 1 and lower in _SINGLE_WORD_HEADERS:
        return True
 
    if len(words) >= 2 and words[-1].lower() in _LABEL_SUFFIXES:
        return True
 
    if len(words) >= 2:
        generic_count = sum(1 for w in words if w.lower() in _GENERIC_WORDS)
        if generic_count == len(words):
            return True
 
    return False


def _is_purely_numeric(val: str) -> bool:
    """Return True if the value is a plain integer/float string (serial numbers, etc.)."""
    try:
        float(val)
        return True
    except ValueError:
        return False


# =========================
# TABULAR HEADER DETECTION
# =========================

FIELD_COL_HEADERS = {
    "loan api fields",
    "fields",
    "field name",
    "field names",
    "parameter",
    "parameter name",
    "input field",
    "input fields",
    "column name",
    "column names",
    "object name",
}

SAMPLE_COL_HEADERS = {
    "sample values",
    "sample value",
    "example",
    "examples",
}

CATEGORY_COL_HEADERS = {
    "field category",
    "category",
    "api name",
    "group",
    "section",
    "module",
}


def find_header_row_and_cols(
    ws, max_scan_rows: int = 10
) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    for row_idx, row in enumerate(
        ws.iter_rows(max_row=max_scan_rows, values_only=True), start=1
    ):
        field_col    = None
        sample_col   = None
        category_col = None
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            norm = str(cell).strip().lower()
            if norm in FIELD_COL_HEADERS and field_col is None:
                field_col = col_idx
            elif norm in SAMPLE_COL_HEADERS and sample_col is None:
                sample_col = col_idx
            elif norm in CATEGORY_COL_HEADERS and category_col is None:
                category_col = col_idx
 
        if field_col is not None:
            return row_idx + 1, field_col, sample_col, category_col
 
    return None, None, None, None


# =========================
# TABULAR SHEET PARSER
# =========================
def parse_tabular_sheet(
    ws, sheet_name: str, global_seen: Optional[set] = None
) -> Optional[List[dict]]:
    """
    Try to parse *ws* as a tabular sheet.
    Returns a list of field dicts, or None if no recognised field header found.

    Dedup rules:
    - If a category column exists: dedup key is (category, field_name)
      → same field name under different categories is kept (e.g. CITY under
        Customer vs Co-applicant).
    - Without a category column: dedup key is field_name alone
      → same field name seen anywhere (across sheets) is kept only once.
    - global_seen is shared across all sheets to enforce cross-sheet dedup.
    """
    if global_seen is None:
        global_seen = set()
 
    data_start, field_col, sample_col, category_col = find_header_row_and_cols(ws)
    if field_col is None:
        return None
 
    fields = []
    local_seen = set()   # within-sheet dedup (for category-aware key)
 
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        val = row[field_col] if field_col < len(row) else None
        if val is None:
            continue
 
        field_name = str(val).strip()
        if not field_name or field_name.lower() == 'nan':
            continue
        if _is_header_label(field_name):
            logger.debug(f"Tabular skip (header label): {field_name!r}")
            continue
        if _is_purely_numeric(field_name):
            logger.debug(f"Tabular skip (numeric): {field_name!r}")
            continue
        if _is_dotted_path(field_name):
            logger.debug(f"Tabular skip (dotted-path): {field_name!r}")
            continue
 
        # Resolve category value
        cat_val = None
        if category_col is not None and category_col < len(row):
            cv = row[category_col]
            if cv is not None:
                cat_val = str(cv).strip()
 
        # Dedup key:
        # - With category column  → (category, field_name): same field in
        #   different categories is a distinct entry.
        # - Without category column → field_name only: one entry per unique
        #   name across the entire workbook.
        if category_col is not None:
            dedup_key = (cat_val, field_name)
            if dedup_key in local_seen:
                continue
            local_seen.add(dedup_key)
        else:
            dedup_key = field_name.lower()
            if dedup_key in global_seen:
                logger.debug(f"Tabular skip (global duplicate): {field_name!r}")
                continue
            global_seen.add(dedup_key)
 
        sample = None
        if sample_col is not None and sample_col < len(row):
            sv = row[sample_col]
            if sv is not None:
                sample = str(sv).strip()
 
        fields.append(
            FieldDefinition(
                field_name=field_name,
                column_category=cat_val if cat_val else sheet_name,
                sample_value=sample,
                source_sheet=sheet_name,
            ).to_dict()
        )
 
    return fields


# =========================
# KV SHEET DETECTION
# =========================
def is_key_value_sheet(ws) -> bool:
    values = []
 
    for i in range(1, min(ws.max_row + 1, 20)):
        v = ws.cell(row=i, column=1).value
        if v:
            values.append(str(v).strip())
 
    if len(values) < 2:
        return False
 
    identifier_ratio = sum(_looks_like_identifier(v) for v in values) / len(values)
    return identifier_ratio > 0.6


# =========================
# PARSE KV SHEET
# =========================
def parse_key_value_sheet(
    ws, sheet_name: str, global_seen: Optional[set] = None
) -> List[dict]:
    """
    Parse a key-value sheet (column A = field names).

    global_seen is shared across all sheets so that the same field name
    appearing in multiple KV sheets (or after a tabular sheet) is kept
    only once.
    """
    if global_seen is None:
        global_seen = set()
 
    fields = []
 
    for i in range(1, ws.max_row + 1):
        val = ws.cell(row=i, column=1).value
 
        if not val:
            continue
 
        field_name = str(val).strip()
 
        if not field_name or field_name.lower() == 'nan':
            continue
        if _is_header_label(field_name):
            logger.debug(f"KV skip (header label): {field_name!r}")
            continue
        if _is_purely_numeric(field_name):
            logger.debug(f"KV skip (numeric): {field_name!r}")
            continue
        if _is_dotted_path(field_name):
            logger.debug(f"KV skip (dotted-path): {field_name!r}")
            continue
 
        dedup_key = field_name.lower()
        if dedup_key in global_seen:
            logger.debug(f"KV skip (global duplicate): {field_name!r}")
            continue
        global_seen.add(dedup_key)
 
        sample = None
        if ws.max_column >= 2:
            v = ws.cell(row=i, column=2).value
            if v:
                sample = str(v).strip()
 
        fields.append(
            FieldDefinition(
                field_name=field_name,
                column_category=sheet_name,
                sample_value=sample,
                source_sheet=sheet_name,
            ).to_dict()
        )
 
    return fields


# =========================
# EXCEL PARSER
# Route: tabular first → KV fallback
# =========================
def parse_excel_input(file_path: str) -> List[dict]:
    wb          = load_workbook(file_path, data_only=True)
    all_fields: List[dict] = []
    global_seen: set       = set()   # shared across all sheets for dedup
 
    for sheet in wb.sheetnames:
        ws = wb[sheet]
 
        if ws.sheet_state != 'visible':
            logger.info(f"{sheet} → Skipped (hidden sheet)")
            continue
 
        # --- Attempt tabular parsing first ---
        result = parse_tabular_sheet(ws, sheet, global_seen)
        if result is not None:
            logger.info(f"{sheet} → Tabular format ({len(result)} fields found)")
            all_fields.extend(result)
            continue
 
        # --- Fallback: KV list ---
        if is_key_value_sheet(ws):
            logger.info(f"{sheet} → KV format")
            all_fields.extend(parse_key_value_sheet(ws, sheet, global_seen))
        else:
            logger.info(f"{sheet} → Skipped (unrecognised format)")
 
    wb.close()
    return all_fields


# =========================
# JSON PARSER
# =========================
def parse_json_input(file_path: str) -> List[dict]:
    raw = Path(file_path).read_text(encoding="utf-8", errors="ignore")

    def _extract_null_fields_from_jsonish(text: str) -> List[dict]:
        """
        Extract partner fields from "JSON-ish" text that might not be valid JSON.

        Rule:
        - Any `"someKey": null` is treated as a partner field named `someKey`.

        column_category:
        - If an object contains `column_category` / `columnCategory`, it applies to
          fields under that object.
        - Otherwise, use the current **section** (nearest parent object key),
          excluding container keys like `loanAccounts` / `loanAccount`, or "API".
        """
        s = text.lstrip("\ufeff")

        # Remove inline notes like: null, - bureau/banking  /  "x", - note
        s = re.sub(
            r"(null|true|false|\"[^\"]*\")\s*,\s*-\s*[^,\r\n]+",
            r"\1",
            s,
            flags=re.I,
        )
        # Drop standalone prose lines that are not JSON (no colon and no brace/bracket)
        s = re.sub(r"(?m)^\s*[^:\{\}\[\]\"]+\s*$", "", s)
        # Remove trailing commas before } or ]
        s = re.sub(r",\s*([}\]])", r"\1", s)

        # Tokenizer: yields ("str", value) and ("sym", one_of_structural) and ("lit", literal)
        def tokens(src: str):
            i = 0
            n = len(src)
            while i < n:
                ch = src[i]
                if ch.isspace():
                    i += 1
                    continue
                if ch in "{}[]:,":
                    yield ("sym", ch)
                    i += 1
                    continue
                if ch == '"':
                    i += 1
                    buf = []
                    while i < n:
                        c = src[i]
                        if c == "\\":
                            if i + 1 < n:
                                buf.append(src[i + 1])
                                i += 2
                                continue
                            i += 1
                            continue
                        if c == '"':
                            i += 1
                            break
                        buf.append(c)
                        i += 1
                    yield ("str", "".join(buf))
                    continue

                # literals: null/true/false (we don't need numbers)
                if src[i : i + 4].lower() == "null":
                    yield ("lit", "null")
                    i += 4
                    continue
                if src[i : i + 4].lower() == "true":
                    yield ("lit", "true")
                    i += 4
                    continue
                if src[i : i + 5].lower() == "false":
                    yield ("lit", "false")
                    i += 5
                    continue

                # Unknown token (e.g. unquoted prose) — skip until next whitespace or structural char
                j = i + 1
                while j < n and (not src[j].isspace()) and src[j] not in "{}[]:,\"":
                    j += 1
                i = j

        # Context stack: each entry is (object_key, category_override)
        stack: List[tuple[Optional[str], Optional[str]]] = []

        _IGNORE_CATEGORY_KEYS = {"loanaccounts", "loanaccount"}

        def _category_from_stack_parts(parts: List[str]) -> str:
            cleaned = [p for p in parts if p and p.strip().lower() not in _IGNORE_CATEGORY_KEYS]
            return (cleaned[-1] if cleaned else "") or "API"

        def current_category() -> str:
            # nearest override wins
            for _k, c in reversed(stack):
                if c:
                    return c
            parts = [k for (k, _c) in stack if k]
            return _category_from_stack_parts(parts)

        out: List[dict] = []
        seen: set[str] = set()

        pending_key: Optional[str] = None
        expecting_value = False
        last_sym: Optional[str] = None

        it = iter(tokens(s))
        for typ, val in it:
            if typ == "sym":
                last_sym = val
                if val == "{":
                    # root object or object value; if it was after a key, push that key
                    if pending_key is not None and expecting_value:
                        stack.append((pending_key, None))
                        pending_key = None
                        expecting_value = False
                    else:
                        stack.append((None, None))
                elif val == "}":
                    if stack:
                        stack.pop()
                    pending_key = None
                    expecting_value = False
                elif val == ":":
                    expecting_value = True
                elif val == ",":
                    pending_key = None
                    expecting_value = False
                continue

            if typ == "str":
                # If this string is a key (next should be :), record it.
                if not expecting_value:
                    pending_key = val
                else:
                    # value is a string
                    if pending_key in ("column_category", "columnCategory", "COLUMN_CATEGORY"):
                        # apply category override to current object (top of stack)
                        if stack:
                            k0, _ = stack.pop()
                            stack.append((k0, val.strip() or None))
                    pending_key = None
                    expecting_value = False
                continue

            if typ == "lit":
                if expecting_value and pending_key:
                    if pending_key in ("column_category", "columnCategory", "COLUMN_CATEGORY"):
                        # ignore non-string category
                        pending_key = None
                        expecting_value = False
                        continue
                    if val.lower() == "null":
                        field_name = (pending_key or "").strip()
                        if field_name and not _is_header_label(field_name) and not _is_purely_numeric(field_name):
                            dedup = f"{current_category().lower()}||{field_name.lower()}"
                            if dedup not in seen:
                                seen.add(dedup)
                                out.append(
                                    FieldDefinition(
                                        field_name=field_name,
                                        column_category=current_category(),
                                        sample_value=None,
                                        source_sheet="API",
                                    ).to_dict()
                                )
                    pending_key = None
                    expecting_value = False
                continue

        return out

    def _sanitize_jsonish(src: str) -> str:
        """
        Make a best-effort attempt to convert 'doc style' JSON into valid JSON:
        - strips inline notes like: null, - bureau/banking
        - removes standalone prose lines (e.g. 'for salaried')
        - removes trailing commas before } or ]
        """
        s = src
        s = re.sub(
            r"(null|true|false|\"[^\"]*\")\s*,\s*-\s*[^,\r\n]+",
            r"\1",
            s,
            flags=re.I,
        )
        s = re.sub(r"(?m)^\s*[^:\{\}\[\]\"]+\s*$", "", s)
        s = re.sub(r",\s*([}\]])", r"\1", s)
        return s

    try:
        # Prefer sanitized parse so we stay on the structured-path logic (better categories)
        data = json.loads(_sanitize_jsonish(raw))
    except Exception:
        # Fallback: still not valid JSON. Extract `"key": null` fields directly.
        return _extract_null_fields_from_jsonish(raw)

    results: List[dict] = []

    _IGNORE_CATEGORY_KEYS = {"loanaccounts", "loanaccount"}

    def _default_category_from_path_parts(path_parts: List[str]) -> str:
        cleaned = [
            p for p in path_parts
            if p and str(p).strip().lower() not in _IGNORE_CATEGORY_KEYS
        ]
        return (str(cleaned[-1]).strip() if cleaned else "") or "API"

    def _category_for(node: Any, path_parts: List[str]) -> str:
        # If node declares a category explicitly, prefer it.
        if isinstance(node, dict):
            for key in ("column_category", "columnCategory", "COLUMN_CATEGORY"):
                v = node.get(key)
                if isinstance(v, str) and v.strip():
                    return v.strip()
        # Otherwise use the nearest parent section key.
        return _default_category_from_path_parts(path_parts)

    def walk(node: Any, path_parts: List[str], inherited_category: str) -> None:
        if isinstance(node, dict):
            # Allow explicit category override at any level
            here_category = _category_for(node, path_parts) or inherited_category
            for k, v in node.items():
                # Don't treat the category metadata key itself as a partner field
                if str(k) in ("column_category", "columnCategory", "COLUMN_CATEGORY"):
                    continue
                walk(v, path_parts + [str(k)], here_category)
            return

        if isinstance(node, list):
            for idx, item in enumerate(node):
                walk(item, path_parts + [str(idx)], inherited_category)
            return

        # Leaf: partner field is any key whose value is null/None
        if node is None and path_parts:
            field_name = str(path_parts[-1]).strip()
            if not field_name:
                return
            if _is_header_label(field_name) or _is_purely_numeric(field_name):
                return
            results.append(
                FieldDefinition(
                    field_name=field_name,
                    column_category=inherited_category or "API",
                    sample_value=None,
                    source_sheet="API",
                ).to_dict()
            )

    walk(data, [], "API")
    return results


# =========================
# MAIN ENTRY
# =========================
def parse_input(file_path: str) -> List[dict]:
    file_path = Path(file_path)
 
    if not file_path.exists():
        raise FileNotFoundError(file_path)
 
    if file_path.suffix in ['.xlsx', '.xls']:
        return parse_excel_input(str(file_path))
 
    if file_path.suffix == '.json':
        return parse_json_input(str(file_path))
 
    raise ValueError("Unsupported file type")


# =========================
# CLI RUN
# =========================
if __name__ == "__main__":
    import sys
 
    file   = sys.argv[1]
    result = parse_input(file)
    print(json.dumps(result, indent=2))